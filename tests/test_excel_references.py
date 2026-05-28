"""
Tests de regresión para tools/excel_generator.py — referencias cruzadas entre hojas.

Protege contra el bug de offsets de fila calculados a mano: la hoja Model y la hoja
Valuation referencian celdas de la hoja Assumptions por número de fila. Si esos números
no cuadran con el layout real, el DCF se calcula sobre celdas equivocadas (p.ej. el WACC
apuntando a "Terminal Value Multiple - Bear", o SG&A leyendo una fila en blanco) y el
modelo entero queda descuadrado sin que nada lo detecte.

La aserción clave NO es "la fórmula no está vacía", sino que la fila destino en
Assumptions tiene la ETIQUETA esperada. Eso es lo que habría cazado el bug original.
"""
import re

import openpyxl
import pytest

from tools.excel_generator import generate_valuation_excel

# Referencia "Assumptions!C38" o "Assumptions!$C$38"
_REF = re.compile(r"Assumptions!\$?([A-Z]+)\$?([0-9]+)")

# Etiqueta de la fila (col A) en Model/Valuation -> substring esperado en la fila
# de Assumptions que esa fila referencia.
_EXPECTED = {
    "Cost of Goods Sold (COGS)": "Gross Margin (Selected)",
    "SG&A Expense": "SG&A % Revenue (Selected)",
    "R&D Expense": "R&D % Revenue (Selected)",
    "Depreciation & Amortization": "D&A % Revenue",
    "Taxes": "Tax Rate",
    "(-) Capital Expenditure": "CapEx % Revenue",
    "WACC": "WACC (Selected)",
    "Terminal Value Multiple": "Terminal Value Multiple (Selected)",
}


def _make_inputs(n_seg):
    """Datos sintéticos mínimos para generar un Excel con n_seg segmentos."""
    years = [2021, 2022, 2023, 2024]
    historical = {
        y: {
            "total_revenue": 100_000e6,
            "cost_of_revenue": 60_000e6,
            "selling_general_admin": 15_000e6,
            "research_development": 5_000e6,
            "depreciation": 4_000e6,
            "interest_expense": 1_000e6,
        }
        for y in years
    }
    segments = [
        {
            "name": f"Segment {i + 1}",
            "revenues": {y: 100_000e6 / n_seg for y in years},
            "pct": 1.0 / n_seg,
        }
        for i in range(n_seg)
    ]
    data = {
        "info": {"longName": "Test Co"},
        "segments": segments,
        "balance_sheet": None,
        "shares_outstanding": 1e9,
        "current_price": 100.0,
    }
    sc = {
        "gross_margin": 0.40,
        "sga_pct": 0.15,
        "rd_pct": 0.05,
        "da_pct": 0.04,
        "capex_pct": 0.05,
        "tax_rate": 0.21,
        "wacc": 0.10,
        "terminal_multiple": 12,
    }
    for y in range(1, 6):
        sc[f"revenue_growth_y{y}"] = 0.05
    metrics = {"_real_scenarios": {"base": dict(sc), "bull": dict(sc), "bear": dict(sc)}}
    return data, historical, metrics


def _build_wb(tmp_path, n_seg):
    data, historical, metrics = _make_inputs(n_seg)
    out = tmp_path / f"T{n_seg}.xlsx"
    generate_valuation_excel("T", data, historical, metrics, output_path=str(out))
    return openpyxl.load_workbook(out, data_only=False)


def _row_refs(ws):
    """Devuelve [(etiqueta_colA, fila_destino_en_Assumptions), ...] para cada fila
    de `ws` que contenga una referencia a la hoja Assumptions."""
    out = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        for col in range(2, 16):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and "Assumptions!" in v:
                m = _REF.search(v)
                if m:
                    out.append((label, int(m.group(2))))
                break
    return out


@pytest.mark.parametrize("n_seg", [1, 2, 5])
def test_drivers_referencian_la_fila_correcta(tmp_path, n_seg):
    """Cada driver económico (GM, SG&A, R&D, D&A, Tax, CapEx, WACC, TV) referencia
    la fila de Assumptions con la etiqueta correcta, para 1, 2 y 5 segmentos."""
    wb = _build_wb(tmp_path, n_seg)
    assumptions = wb["Assumptions"]

    checked = 0
    for sheet_name in ("Model", "Valuation"):
        ws = wb[sheet_name]
        for label, target_row in _row_refs(ws):
            if label not in _EXPECTED:
                continue
            got = assumptions.cell(row=target_row, column=1).value
            expected = _EXPECTED[label]
            assert got is not None and expected in str(got), (
                f"[n_seg={n_seg}] {sheet_name}!{label!r} referencia Assumptions fila "
                f"{target_row} (etiqueta {got!r}), se esperaba {expected!r}"
            )
            checked += 1

    # Todos los drivers del mapa deben haberse comprobado (si no, el layout cambió
    # de nombre y este test dejó de proteger lo que cree proteger).
    assert checked >= len(_EXPECTED), (
        f"Solo se comprobaron {checked} drivers de {len(_EXPECTED)} esperados; "
        "¿cambiaron las etiquetas de fila?"
    )


@pytest.mark.parametrize("n_seg", [1, 2, 5])
def test_ninguna_referencia_cae_en_fila_vacia_o_cabecera(tmp_path, n_seg):
    """Ninguna referencia a Assumptions debe aterrizar en una fila en blanco o en una
    cabecera de sección — ese era el síntoma del bug (SG&A leía una celda vacía, Tax
    leía la cabecera 'Valuation Assumptions')."""
    wb = _build_wb(tmp_path, n_seg)
    assumptions = wb["Assumptions"]
    headers = {"Operating Assumptions", "Valuation Assumptions", "Profit & Loss",
               "Revenue by Segment", "Cash Flow"}

    for sheet_name in ("Model", "Valuation"):
        for label, target_row in _row_refs(wb[sheet_name]):
            target_label = assumptions.cell(row=target_row, column=1).value
            # El selector de escenario (Assumptions!G2) es una referencia legítima a
            # una fila sin etiqueta en col A; se ignora.
            if target_row == 2:
                continue
            assert target_label not in (None, ""), (
                f"[n_seg={n_seg}] {sheet_name}!{label!r} referencia la fila vacía "
                f"{target_row} de Assumptions"
            )
            assert target_label not in headers, (
                f"[n_seg={n_seg}] {sheet_name}!{label!r} referencia la cabecera "
                f"{target_label!r} (fila {target_row}) en vez de un dato"
            )

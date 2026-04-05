"""
Genera el Excel de valoracion completo.
Replica la estructura del modelo Netflix con 4 hojas:
Assumptions, Financial Statements, Model, Valuation.

Todas las proyecciones usan formulas reales de Excel (no valores hardcodeados).
Selector de escenario con OFFSET para cambiar entre Base/Bull/Bear.
Tabla de sensibilidad WACC x Terminal Value Multiple.
"""

import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# ESTILOS
# ============================================================
HEADER_FILL = PatternFill(start_color="FF273A4F", end_color="FF273A4F", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
INPUT_FILL = PatternFill(start_color="FFB2C3BC", end_color="FFB2C3BC", fill_type="solid")
INPUT_FONT = Font(name="Calibri", size=10, color="FF0000FF")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
PCT_FORMAT = '0.0%'
NUM_FORMAT = '#,##0'
NUM_FORMAT_M = '#,##0.0'
CURRENCY_FORMAT = '$#,##0.00'
THIN_BORDER = Border(bottom=Side(style='thin', color='FFD0D0D0'))


def _set_header(ws, row, col_start, col_end, text):
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = HEADER_FILL
        ws.cell(row=row, column=c).font = HEADER_FONT
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")


def _set_input(ws, row, col, value, is_pct=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    if is_pct:
        cell.number_format = PCT_FORMAT
    elif isinstance(value, (int, float)):
        cell.number_format = NUM_FORMAT


def _set_label(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD_FONT if bold else NORMAL_FONT


def _set_value(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = NORMAL_FONT
    if fmt:
        cell.number_format = fmt


def _col(n):
    return get_column_letter(n)


# ============================================================
# FUNCION PRINCIPAL
# ============================================================

def generate_valuation_excel(ticker: str, data: dict, historical: dict,
                              metrics: dict, output_path: str = None) -> str:
    """
    Genera el Excel de valoración con datos históricos y template DCF.

    Los 3 escenarios (Base/Bull/Bear) se rellenan con valores de referencia
    históricos como punto de partida. El usuario o Opus los edita manualmente.

    Args:
        ticker: Ticker de la empresa
        data: Datos de get_company_data()
        historical: Datos de extract_historical_data()
        metrics: Datos de extract_metrics() (márgenes, growth, detecciones)
        output_path: Path de salida
    """
    if output_path is None:
        output_path = os.path.join(ticker, f"{ticker}_modelo_valoracion.xlsx")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    print(f"  [excel] Generando modelo de valoracion para {ticker}...")

    # Usar escenarios reales si los proporcionó Opus, si no placeholders de referencia
    real = metrics.get("_real_scenarios") if isinstance(metrics, dict) else None
    if real and all(k in real for k in ("bear", "base", "bull")):
        scenarios = real
        print(f"    → Usando escenarios reales de Opus")
    else:
        avg = metrics.get("avg_margins", {})
        avg_growth = metrics.get("avg_growth", 0.05) or 0.05
        ref_scenario = {
            "gross_margin": avg.get("gross_margin", 0.40),
            "sga_pct": avg.get("sga_pct", 0.15),
            "rd_pct": avg.get("rd_pct", 0.0),
            "da_pct": avg.get("da_pct", 0.03),
            "capex_pct": avg.get("capex_pct", 0.04),
            "tax_rate": avg.get("tax_rate", 0.21),
            "wacc": 0.10,
            "terminal_multiple": 12,
        }
        for y in range(1, 6):
            ref_scenario[f"revenue_growth_y{y}"] = round(avg_growth * (1 - 0.1 * (y - 1)), 4)
        scenarios = {"base": ref_scenario, "bull": dict(ref_scenario), "bear": dict(ref_scenario)}
        print(f"    → Usando márgenes históricos como referencia (editar en Excel)")

    wb = Workbook()
    info = data["info"]
    company_name = info.get("longName", ticker)
    segments = data["segments"]
    sorted_years = sorted(historical.keys(), reverse=False)

    last_hist_year = max(sorted_years) if sorted_years else 2024
    proj_years = list(range(last_hist_year + 1, last_hist_year + 6))
    all_years = sorted_years + proj_years

    n_hist = len(sorted_years)
    n_proj = len(proj_years)

    first_data_col = 3
    hist_cols = list(range(first_data_col, first_data_col + n_hist))
    proj_cols = list(range(first_data_col + n_hist, first_data_col + n_hist + n_proj))
    all_cols = hist_cols + proj_cols

    # HOJA 1: ASSUMPTIONS
    ws_a = wb.active
    ws_a.title = "Assumptions"
    _build_assumptions_sheet(ws_a, ticker, company_name, segments, scenarios,
                             historical, sorted_years, proj_years, all_years,
                             first_data_col, hist_cols, proj_cols, all_cols)

    # HOJA 2: FINANCIAL STATEMENTS
    ws_fs = wb.create_sheet("Financial Statements")
    _build_financial_statements_sheet(ws_fs, ticker, company_name, historical,
                                       sorted_years, first_data_col, hist_cols)

    # HOJA 3: MODEL
    ws_m = wb.create_sheet("Model")
    _build_model_sheet(ws_m, ticker, company_name, segments, scenarios,
                       historical, sorted_years, proj_years, all_years,
                       first_data_col, hist_cols, proj_cols, all_cols, data)

    # HOJA 4: VALUATION
    ws_v = wb.create_sheet("Valuation")
    _build_valuation_sheet(ws_v, ticker, company_name, scenarios, data,
                           proj_years, proj_cols, first_data_col, n_hist,
                           metrics=metrics)

    # Ajustar anchos
    for ws in [ws_a, ws_fs, ws_m, ws_v]:
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 3
        for col_idx in all_cols:
            ws.column_dimensions[_col(col_idx)].width = 15
        ws.sheet_properties.tabColor = "273A4F"

    wb.save(output_path)
    print(f"    Excel guardado: {output_path}")
    return output_path


# ============================================================
# ASSUMPTIONS SHEET
# ============================================================

def _build_assumptions_sheet(ws, ticker, company_name, segments, scenarios,
                              historical, sorted_years, proj_years, all_years,
                              first_data_col, hist_cols, proj_cols, all_cols):

    _set_header(ws, 1, 1, max(all_cols), f"{company_name} ({ticker}) - Assumptions")

    # Selector de escenario en G2
    ws.cell(row=2, column=1, value="Scenario Selector:").font = BOLD_FONT
    cell_scenario = ws.cell(row=2, column=7, value=1)
    cell_scenario.font = INPUT_FONT
    cell_scenario.fill = INPUT_FILL
    ws.cell(row=2, column=8, value="1=Base, 2=Bull, 3=Bear").font = NORMAL_FONT

    # Headers de anios
    row = 4
    for i, year in enumerate(all_years):
        col = first_data_col + i
        cell = ws.cell(row=row, column=col, value=f"FY{year}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # REVENUE ASSUMPTIONS POR SEGMENTO
    row = 6
    base_sc = scenarios["base"]
    bull_sc = scenarios["bull"]
    bear_sc = scenarios["bear"]

    for seg_idx, segment in enumerate(segments):
        seg_name = segment["name"]
        _set_header(ws, row, 1, max(all_cols), seg_name)
        row += 1

        # Revenue historico
        _set_label(ws, row, 1, "  Revenue ($M)", bold=True)
        revenue_row = row
        seg_revenues = segment.get("revenues", {})
        for i, year in enumerate(sorted_years):
            col = hist_cols[i]
            rev = seg_revenues.get(year, 0)
            if rev == 0 and historical.get(year, {}).get("total_revenue"):
                pct = segment.get("pct", 1.0)
                rev = historical[year]["total_revenue"] * pct
            _set_value(ws, row, col, rev / 1e6 if rev > 1e3 else rev, NUM_FORMAT_M)
        row += 1

        # Growth rates (3 escenarios)
        _set_label(ws, row, 1, "  Revenue Growth - Base")
        base_growth_row = row
        for i in range(len(proj_years)):
            _set_input(ws, row, proj_cols[i], base_sc[f"revenue_growth_y{i+1}"], is_pct=True)
        row += 1

        _set_label(ws, row, 1, "  Revenue Growth - Bull")
        for i in range(len(proj_years)):
            _set_input(ws, row, proj_cols[i], bull_sc[f"revenue_growth_y{i+1}"], is_pct=True)
        row += 1

        _set_label(ws, row, 1, "  Revenue Growth - Bear")
        for i in range(len(proj_years)):
            _set_input(ws, row, proj_cols[i], bear_sc[f"revenue_growth_y{i+1}"], is_pct=True)
        row += 1

        # Selected Growth (OFFSET)
        _set_label(ws, row, 1, "  Revenue Growth (Selected)", bold=True)
        selected_growth_row = row
        for i in range(len(proj_years)):
            col = proj_cols[i]
            c_letter = _col(col)
            formula = f'=OFFSET({c_letter}{base_growth_row},$G$2-1,0)'
            ws.cell(row=row, column=col, value=formula).number_format = PCT_FORMAT
        row += 1

        # Revenue proyectado
        _set_label(ws, row, 1, "  Revenue Projected ($M)", bold=True)
        rev_proj_row = row
        for i in range(len(proj_years)):
            col = proj_cols[i]
            c_letter = _col(col)
            prev_col_letter = _col(col - 1)
            prev_ref = revenue_row if i == 0 else rev_proj_row
            formula = f'={prev_col_letter}{prev_ref}*(1+{c_letter}{selected_growth_row})'
            ws.cell(row=row, column=col, value=formula).number_format = NUM_FORMAT_M
        row += 2

    # OPERATING ASSUMPTIONS
    _set_header(ws, row, 1, max(all_cols), "Operating Assumptions")
    row += 1

    # Gross Margin (3 escenarios + Selected)
    _set_label(ws, row, 1, "Gross Margin - Base")
    gm_base_row = row
    for col in proj_cols:
        _set_input(ws, row, col, base_sc["gross_margin"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "Gross Margin - Bull")
    for col in proj_cols:
        _set_input(ws, row, col, bull_sc["gross_margin"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "Gross Margin - Bear")
    for col in proj_cols:
        _set_input(ws, row, col, bear_sc["gross_margin"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "Gross Margin (Selected)", bold=True)
    for col in proj_cols:
        c = _col(col)
        ws.cell(row=row, column=col, value=f'=OFFSET({c}{gm_base_row},$G$2-1,0)').number_format = PCT_FORMAT
    row += 2

    # SG&A
    _set_label(ws, row, 1, "SG&A % Revenue - Base")
    sga_base_row = row
    for col in proj_cols:
        _set_input(ws, row, col, base_sc["sga_pct"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "SG&A % Revenue - Bull")
    for col in proj_cols:
        _set_input(ws, row, col, bull_sc["sga_pct"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "SG&A % Revenue - Bear")
    for col in proj_cols:
        _set_input(ws, row, col, bear_sc["sga_pct"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "SG&A % Revenue (Selected)", bold=True)
    for col in proj_cols:
        c = _col(col)
        ws.cell(row=row, column=col, value=f'=OFFSET({c}{sga_base_row},$G$2-1,0)').number_format = PCT_FORMAT
    row += 2

    # R&D
    _set_label(ws, row, 1, "R&D % Revenue - Base")
    rd_base_row = row
    for col in proj_cols:
        _set_input(ws, row, col, base_sc["rd_pct"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "R&D % Revenue - Bull")
    for col in proj_cols:
        _set_input(ws, row, col, bull_sc["rd_pct"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "R&D % Revenue - Bear")
    for col in proj_cols:
        _set_input(ws, row, col, bear_sc["rd_pct"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "R&D % Revenue (Selected)", bold=True)
    for col in proj_cols:
        c = _col(col)
        ws.cell(row=row, column=col, value=f'=OFFSET({c}{rd_base_row},$G$2-1,0)').number_format = PCT_FORMAT
    row += 2

    # D&A, CapEx, Tax (sin escenarios, solo base)
    _set_label(ws, row, 1, "D&A % Revenue")
    for col in proj_cols:
        _set_input(ws, row, col, base_sc["da_pct"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "CapEx % Revenue")
    for col in proj_cols:
        _set_input(ws, row, col, base_sc["capex_pct"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "Tax Rate")
    for col in proj_cols:
        _set_input(ws, row, col, base_sc["tax_rate"], is_pct=True)
    row += 2

    # WACC & TERMINAL VALUE
    _set_header(ws, row, 1, max(all_cols), "Valuation Assumptions")
    row += 1

    _set_label(ws, row, 1, "WACC - Base")
    wacc_base_row = row
    _set_input(ws, row, first_data_col, base_sc["wacc"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "WACC - Bull")
    _set_input(ws, row, first_data_col, bull_sc["wacc"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "WACC - Bear")
    _set_input(ws, row, first_data_col, bear_sc["wacc"], is_pct=True)
    row += 1
    _set_label(ws, row, 1, "WACC (Selected)", bold=True)
    c = _col(first_data_col)
    ws.cell(row=row, column=first_data_col,
            value=f'=OFFSET({c}{wacc_base_row},$G$2-1,0)').number_format = PCT_FORMAT
    row += 2

    _set_label(ws, row, 1, "Terminal Value Multiple - Base")
    tv_base_row = row
    _set_input(ws, row, first_data_col, base_sc["terminal_multiple"])
    ws.cell(row=row, column=first_data_col).number_format = '0.0"x"'
    row += 1
    _set_label(ws, row, 1, "Terminal Value Multiple - Bull")
    _set_input(ws, row, first_data_col, bull_sc["terminal_multiple"])
    ws.cell(row=row, column=first_data_col).number_format = '0.0"x"'
    row += 1
    _set_label(ws, row, 1, "Terminal Value Multiple - Bear")
    _set_input(ws, row, first_data_col, bear_sc["terminal_multiple"])
    ws.cell(row=row, column=first_data_col).number_format = '0.0"x"'
    row += 1
    _set_label(ws, row, 1, "Terminal Value Multiple (Selected)", bold=True)
    c = _col(first_data_col)
    ws.cell(row=row, column=first_data_col,
            value=f'=OFFSET({c}{tv_base_row},$G$2-1,0)').number_format = '0.0"x"'


# ============================================================
# FINANCIAL STATEMENTS SHEET
# ============================================================

def _build_financial_statements_sheet(ws, ticker, company_name, historical,
                                        sorted_years, first_data_col, hist_cols):

    max_col = max(hist_cols) if hist_cols else 5
    _set_header(ws, 1, 1, max_col, f"{company_name} ({ticker}) - Historical Financial Statements")

    row = 3
    _set_label(ws, row, 1, "($ in millions)")
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        cell = ws.cell(row=row, column=col, value=f"FY{year}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # INCOME STATEMENT
    row = 5
    _set_header(ws, row, 1, max_col, "Income Statement")
    row += 1

    is_items = [
        ("Total Revenue", "total_revenue"),
        ("Cost of Revenue", "cost_of_revenue"),
        ("Gross Profit", "gross_profit"),
        ("", None),
        ("R&D Expense", "research_development"),
        ("SG&A Expense", "selling_general_admin"),
        ("Operating Income (EBIT)", "operating_income"),
        ("", None),
        ("Interest Expense", "interest_expense"),
        ("Tax Provision", "tax_provision"),
        ("Net Income", "net_income"),
        ("", None),
        ("EBITDA", "ebitda"),
        ("Diluted EPS", "diluted_eps"),
        ("D&A", "depreciation"),
    ]

    margin_items = {
        "Gross Profit": ("gross_profit", "total_revenue", "Gross Margin"),
        "Operating Income (EBIT)": ("operating_income", "total_revenue", "EBIT Margin"),
        "Net Income": ("net_income", "total_revenue", "Net Margin"),
        "EBITDA": ("ebitda", "total_revenue", "EBITDA Margin"),
    }

    for label, key in is_items:
        if key is None:
            row += 1
            continue

        is_bold = label in ["Total Revenue", "Gross Profit", "Operating Income (EBIT)",
                            "Net Income", "EBITDA"]
        _set_label(ws, row, 1, label, bold=is_bold)
        for i, year in enumerate(sorted_years):
            col = hist_cols[i]
            val = historical.get(year, {}).get(key, 0)
            display_val = val / 1e6 if abs(val) > 1e3 else val
            fmt = CURRENCY_FORMAT if key == "diluted_eps" else NUM_FORMAT_M
            _set_value(ws, row, col, display_val, fmt)
        row += 1

        if label in margin_items:
            num_key, den_key, margin_label = margin_items[label]
            _set_label(ws, row, 1, f"  {margin_label}")
            for i, year in enumerate(sorted_years):
                col = hist_cols[i]
                num = historical.get(year, {}).get(num_key, 0)
                den = historical.get(year, {}).get(den_key, 0)
                margin = num / den if den != 0 else 0
                _set_value(ws, row, col, margin, PCT_FORMAT)
            row += 1

    row += 2

    # BALANCE SHEET
    _set_header(ws, row, 1, max_col, "Balance Sheet")
    row += 1
    bs_items = [
        ("Cash & Equivalents", "cash"),
        ("Current Assets", "current_assets"),
        ("Net PP&E", "net_ppe"),
        ("Goodwill", "goodwill"),
        ("Total Assets", "total_assets"),
        ("", None),
        ("Current Liabilities", "current_liabilities"),
        ("Total Debt", "total_debt"),
        ("Total Liabilities", "total_liabilities"),
        ("", None),
        ("Total Equity", "total_equity"),
    ]
    for label, key in bs_items:
        if key is None:
            row += 1
            continue
        is_bold = label in ["Total Assets", "Total Liabilities", "Total Equity"]
        _set_label(ws, row, 1, label, bold=is_bold)
        for i, year in enumerate(sorted_years):
            col = hist_cols[i]
            val = historical.get(year, {}).get(key, 0)
            _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)
        row += 1

    row += 2

    # CASH FLOW
    _set_header(ws, row, 1, max_col, "Cash Flow Statement")
    row += 1
    cf_items = [
        ("Operating Cash Flow", "operating_cashflow"),
        ("Capital Expenditure", "capex"),
        ("Free Cash Flow", "free_cashflow"),
        ("", None),
        ("D&A", "depreciation_cf"),
        ("Stock-Based Compensation", "stock_based_comp"),
        ("Change in Working Capital", "change_working_capital"),
    ]
    for label, key in cf_items:
        if key is None:
            row += 1
            continue
        is_bold = label in ["Operating Cash Flow", "Free Cash Flow"]
        _set_label(ws, row, 1, label, bold=is_bold)
        for i, year in enumerate(sorted_years):
            col = hist_cols[i]
            val = historical.get(year, {}).get(key, 0)
            _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)
        row += 1


# ============================================================
# MODEL SHEET
# ============================================================

def _build_model_sheet(ws, ticker, company_name, segments, scenarios,
                       historical, sorted_years, proj_years, all_years,
                       first_data_col, hist_cols, proj_cols, all_cols, data):

    last_col = max(all_cols)
    _set_header(ws, 1, 1, last_col, f"{company_name} ({ticker}) - Financial Model")

    ws.cell(row=2, column=1, value="Scenario:").font = BOLD_FONT
    ws.cell(row=2, column=3, value="=Assumptions!G2").font = INPUT_FONT

    # Year headers
    row = 4
    _set_label(ws, row, 1, "($M)")
    for i, year in enumerate(all_years):
        col = first_data_col + i
        cell = ws.cell(row=row, column=col, value=f"FY{year}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        if year in proj_years:
            ws.cell(row=row + 1, column=col, value="Proj.").font = Font(
                name="Calibri", size=8, italic=True, color="FF888888")

    # REVENUE BY SEGMENT
    row = 7
    _set_header(ws, row, 1, last_col, "Revenue by Segment")
    row += 1

    segment_rev_rows = []
    for seg_idx, segment in enumerate(segments):
        _set_label(ws, row, 1, f"  {segment['name']}", bold=True)
        seg_rev_row = row

        seg_revenues = segment.get("revenues", {})
        for i, year in enumerate(sorted_years):
            col = hist_cols[i]
            rev = seg_revenues.get(year, 0)
            if rev == 0 and historical.get(year, {}).get("total_revenue"):
                pct = segment.get("pct", 1.0)
                rev = historical[year]["total_revenue"] * pct
            _set_value(ws, row, col, rev / 1e6 if rev > 1e3 else rev, NUM_FORMAT_M)

        assumptions_rev_proj_row = 6 + seg_idx * 8 + 6
        for i in range(len(proj_years)):
            col = proj_cols[i]
            c_letter = _col(col)
            formula = f"=Assumptions!{c_letter}{assumptions_rev_proj_row}"
            ws.cell(row=row, column=col, value=formula).number_format = NUM_FORMAT_M

        segment_rev_rows.append(seg_rev_row)
        row += 1

        # YoY Growth
        _set_label(ws, row, 1, f"    YoY Growth")
        for i in range(1, len(all_years)):
            col = first_data_col + i
            prev_col = _col(col - 1)
            curr_col = _col(col)
            formula = f'=IFERROR({curr_col}{seg_rev_row}/{prev_col}{seg_rev_row}-1,"-")'
            ws.cell(row=row, column=col, value=formula).number_format = PCT_FORMAT
        row += 1

    # Total Revenue
    row += 1
    _set_label(ws, row, 1, "Total Revenue", bold=True)
    total_rev_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        if len(segment_rev_rows) == 1:
            formula = f"={c}{segment_rev_rows[0]}"
        else:
            refs = "+".join([f"{c}{r}" for r in segment_rev_rows])
            formula = f"={refs}"
        ws.cell(row=row, column=col_idx, value=formula).number_format = NUM_FORMAT_M
    row += 1

    _set_label(ws, row, 1, "  Revenue Growth %")
    for i in range(1, len(all_years)):
        col = first_data_col + i
        prev = _col(col - 1)
        curr = _col(col)
        formula = f'=IFERROR({curr}{total_rev_row}/{prev}{total_rev_row}-1,"-")'
        ws.cell(row=row, column=col, value=formula).number_format = PCT_FORMAT
    row += 2

    # P&L
    _set_header(ws, row, 1, last_col, "Profit & Loss")
    row += 1

    _set_label(ws, row, 1, "Revenue", bold=True)
    pl_rev_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f"={c}{total_rev_row}").number_format = NUM_FORMAT_M
    row += 1

    # COGS
    _set_label(ws, row, 1, "Cost of Goods Sold (COGS)")
    cogs_row = row
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        val = historical.get(year, {}).get("cost_of_revenue", 0)
        _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)

    n_seg = len(segments)
    gm_selected_row_assumptions = 6 + n_seg * 8 + 4
    for i in range(len(proj_years)):
        col = proj_cols[i]
        c = _col(col)
        formula = f"=-{c}{pl_rev_row}*(1-Assumptions!{c}{gm_selected_row_assumptions})"
        ws.cell(row=row, column=col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # Gross Profit
    _set_label(ws, row, 1, "Gross Profit", bold=True)
    gp_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f"={c}{pl_rev_row}+{c}{cogs_row}").number_format = NUM_FORMAT_M
    row += 1
    _set_label(ws, row, 1, "  Gross Margin %")
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f'=IFERROR({c}{gp_row}/{c}{pl_rev_row},"-")').number_format = PCT_FORMAT
    row += 2

    # SG&A
    _set_label(ws, row, 1, "SG&A Expense")
    sga_row = row
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        val = historical.get(year, {}).get("selling_general_admin", 0)
        _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)
    sga_selected_row_assumptions = gm_selected_row_assumptions + 6
    for i in range(len(proj_years)):
        col = proj_cols[i]
        c = _col(col)
        formula = f"=-{c}{pl_rev_row}*Assumptions!{c}{sga_selected_row_assumptions}"
        ws.cell(row=row, column=col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # R&D
    _set_label(ws, row, 1, "R&D Expense")
    rd_row = row
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        val = historical.get(year, {}).get("research_development", 0)
        _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)
    rd_selected_row_assumptions = sga_selected_row_assumptions + 6
    for i in range(len(proj_years)):
        col = proj_cols[i]
        c = _col(col)
        formula = f"=-{c}{pl_rev_row}*Assumptions!{c}{rd_selected_row_assumptions}"
        ws.cell(row=row, column=col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # D&A
    _set_label(ws, row, 1, "Depreciation & Amortization")
    da_row = row
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        val = historical.get(year, {}).get("depreciation", 0) or historical.get(year, {}).get("depreciation_cf", 0)
        _set_value(ws, row, col, abs(val) / 1e6 if abs(val) > 1e3 else abs(val), NUM_FORMAT_M)
    da_row_assumptions = rd_selected_row_assumptions + 2
    for i in range(len(proj_years)):
        col = proj_cols[i]
        c = _col(col)
        formula = f"=-{c}{pl_rev_row}*Assumptions!{c}{da_row_assumptions}"
        ws.cell(row=row, column=col, value=formula).number_format = NUM_FORMAT_M
    row += 2

    # EBITDA
    _set_label(ws, row, 1, "EBITDA", bold=True)
    ebitda_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f"={c}{gp_row}+{c}{sga_row}+{c}{rd_row}").number_format = NUM_FORMAT_M
    row += 1
    _set_label(ws, row, 1, "  EBITDA Margin %")
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f'=IFERROR({c}{ebitda_row}/{c}{pl_rev_row},"-")').number_format = PCT_FORMAT
    row += 1

    # EBIT
    _set_label(ws, row, 1, "EBIT", bold=True)
    ebit_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f"={c}{ebitda_row}+{c}{da_row}").number_format = NUM_FORMAT_M
    row += 1
    _set_label(ws, row, 1, "  EBIT Margin %")
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f'=IFERROR({c}{ebit_row}/{c}{pl_rev_row},"-")').number_format = PCT_FORMAT
    row += 2

    # Interest Expense
    _set_label(ws, row, 1, "Interest Expense")
    interest_row = row
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        val = historical.get(year, {}).get("interest_expense", 0)
        _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)
    if sorted_years:
        last_interest = historical.get(sorted_years[-1], {}).get("interest_expense", 0)
        last_interest_m = last_interest / 1e6 if abs(last_interest) > 1e3 else last_interest
        for col in proj_cols:
            _set_value(ws, row, col, last_interest_m, NUM_FORMAT_M)
    row += 1

    # EBT
    _set_label(ws, row, 1, "Earnings Before Tax (EBT)", bold=True)
    ebt_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f"={c}{ebit_row}+{c}{interest_row}").number_format = NUM_FORMAT_M
    row += 1

    # Taxes
    _set_label(ws, row, 1, "Taxes")
    taxes_row = row
    tax_rate_assumptions = da_row_assumptions + 2
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        val = historical.get(year, {}).get("tax_provision", 0)
        _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)
    for i in range(len(proj_years)):
        col = proj_cols[i]
        c = _col(col)
        formula = f"=-ABS({c}{ebt_row})*Assumptions!{c}{tax_rate_assumptions}"
        ws.cell(row=row, column=col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # Net Income
    _set_label(ws, row, 1, "Net Income", bold=True)
    ni_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f"={c}{ebt_row}+{c}{taxes_row}").number_format = NUM_FORMAT_M
    row += 1
    _set_label(ws, row, 1, "  Net Margin %")
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f'=IFERROR({c}{ni_row}/{c}{pl_rev_row},"-")').number_format = PCT_FORMAT
    row += 3

    # CASH FLOW
    _set_header(ws, row, 1, last_col, "Cash Flow")
    row += 1

    _set_label(ws, row, 1, "EBITDA", bold=True)
    cf_ebitda_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f"={c}{ebitda_row}").number_format = NUM_FORMAT_M
    row += 1

    _set_label(ws, row, 1, "(-) Taxes")
    cf_taxes_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx, value=f"={c}{taxes_row}").number_format = NUM_FORMAT_M
    row += 1

    _set_label(ws, row, 1, "(+/-) Change in Working Capital")
    cf_wc_row = row
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        val = historical.get(year, {}).get("change_working_capital", 0)
        _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)
    for col in proj_cols:
        _set_value(ws, row, col, 0, NUM_FORMAT_M)
    row += 1

    _set_label(ws, row, 1, "Operating Cash Flow", bold=True)
    ocf_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx,
                value=f"={c}{cf_ebitda_row}+{c}{cf_taxes_row}+{c}{cf_wc_row}").number_format = NUM_FORMAT_M
    row += 2

    # CapEx
    _set_label(ws, row, 1, "(-) Capital Expenditure")
    capex_row = row
    capex_row_assumptions = da_row_assumptions + 1
    for i, year in enumerate(sorted_years):
        col = hist_cols[i]
        val = historical.get(year, {}).get("capex", 0)
        _set_value(ws, row, col, val / 1e6 if abs(val) > 1e3 else val, NUM_FORMAT_M)
    for i in range(len(proj_years)):
        col = proj_cols[i]
        c = _col(col)
        formula = f"=-ABS({c}{pl_rev_row})*Assumptions!{c}{capex_row_assumptions}"
        ws.cell(row=row, column=col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # UFCF
    _set_label(ws, row, 1, "Unlevered Free Cash Flow (UFCF)", bold=True)
    ufcf_row = row
    for col_idx in all_cols:
        c = _col(col_idx)
        ws.cell(row=row, column=col_idx,
                value=f"={c}{ocf_row}+{c}{capex_row}").number_format = NUM_FORMAT_M
    row += 1

    # Store row references for Valuation sheet
    ref_row = 200
    ws.cell(row=ref_row, column=1, value="__REF_EBITDA_ROW").font = Font(color="FFFFFFFF")
    ws.cell(row=ref_row, column=2, value=ebitda_row)
    ws.cell(row=ref_row + 1, column=1, value="__REF_TAXES_ROW").font = Font(color="FFFFFFFF")
    ws.cell(row=ref_row + 1, column=2, value=taxes_row)
    ws.cell(row=ref_row + 2, column=1, value="__REF_CAPEX_ROW").font = Font(color="FFFFFFFF")
    ws.cell(row=ref_row + 2, column=2, value=capex_row)
    ws.cell(row=ref_row + 3, column=1, value="__REF_UFCF_ROW").font = Font(color="FFFFFFFF")
    ws.cell(row=ref_row + 3, column=2, value=ufcf_row)
    ws.cell(row=ref_row + 4, column=1, value="__REF_WC_ROW").font = Font(color="FFFFFFFF")
    ws.cell(row=ref_row + 4, column=2, value=cf_wc_row)
    ws.cell(row=ref_row + 5, column=1, value="__REF_NI_ROW").font = Font(color="FFFFFFFF")
    ws.cell(row=ref_row + 5, column=2, value=ni_row)
    ws.cell(row=ref_row + 6, column=1, value="__REF_TOTAL_REV_ROW").font = Font(color="FFFFFFFF")
    ws.cell(row=ref_row + 6, column=2, value=total_rev_row)


# ============================================================
# VALUATION SHEET
# ============================================================

def _build_valuation_sheet(ws, ticker, company_name, scenarios, data,
                           proj_years, proj_cols, first_data_col, n_hist,
                           metrics=None):

    info = data["info"]
    shares = data["shares_outstanding"]
    current_price = data["current_price"]
    base_sc = scenarios["base"]

    last_proj_col = max(proj_cols) if proj_cols else 10
    n_proj = len(proj_years)

    _set_header(ws, 1, 1, last_proj_col, f"{company_name} ({ticker}) - DCF Valuation")

    ws.cell(row=2, column=1, value="Scenario:").font = BOLD_FONT
    ws.cell(row=2, column=3, value="=Assumptions!G2").font = INPUT_FONT
    ws.cell(row=2, column=5, value="Current Price:").font = BOLD_FONT
    ws.cell(row=2, column=6, value=current_price).font = BOLD_FONT
    ws.cell(row=2, column=6).number_format = CURRENCY_FORMAT

    # Year headers
    row = 4
    _set_label(ws, row, 1, "Year")
    for i, year in enumerate(proj_years):
        col = proj_cols[i]
        cell = ws.cell(row=row, column=col, value=year)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row = 5
    _set_label(ws, row, 1, "Year #")
    for i in range(n_proj):
        ws.cell(row=row, column=proj_cols[i], value=i + 1)

    # DCF INPUTS
    row = 7
    _set_header(ws, row, 1, last_proj_col, "DCF Inputs (from Model)")
    row += 1

    # EBITDA
    _set_label(ws, row, 1, "EBITDA ($M)")
    ebitda_dcf_row = row
    for i in range(n_proj):
        model_col = _col(first_data_col + n_hist + i)
        val_col = proj_cols[i]
        formula = f'=INDIRECT("Model!"&"{model_col}"&Model!B200)'
        ws.cell(row=row, column=val_col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # Taxes
    _set_label(ws, row, 1, "(-) Taxes ($M)")
    taxes_dcf_row = row
    for i in range(n_proj):
        model_col = _col(first_data_col + n_hist + i)
        val_col = proj_cols[i]
        formula = f'=INDIRECT("Model!"&"{model_col}"&Model!B201)'
        ws.cell(row=row, column=val_col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # WC
    _set_label(ws, row, 1, "(+/-) Change in WC ($M)")
    wc_dcf_row = row
    for i in range(n_proj):
        model_col = _col(first_data_col + n_hist + i)
        val_col = proj_cols[i]
        formula = f'=INDIRECT("Model!"&"{model_col}"&Model!B204)'
        ws.cell(row=row, column=val_col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # CapEx
    _set_label(ws, row, 1, "(-) CapEx ($M)")
    capex_dcf_row = row
    for i in range(n_proj):
        model_col = _col(first_data_col + n_hist + i)
        val_col = proj_cols[i]
        formula = f'=INDIRECT("Model!"&"{model_col}"&Model!B202)'
        ws.cell(row=row, column=val_col, value=formula).number_format = NUM_FORMAT_M
    row += 1

    # UFCF
    _set_label(ws, row, 1, "Unlevered Free Cash Flow ($M)", bold=True)
    ufcf_dcf_row = row
    for i in range(n_proj):
        val_col = proj_cols[i]
        c = _col(val_col)
        formula = f"={c}{ebitda_dcf_row}+{c}{taxes_dcf_row}+{c}{wc_dcf_row}+{c}{capex_dcf_row}"
        ws.cell(row=row, column=val_col, value=formula).number_format = NUM_FORMAT_M
    row += 2

    # WACC & DISCOUNTING
    _set_header(ws, row, 1, last_proj_col, "Discounted Cash Flow")
    row += 1

    _set_label(ws, row, 1, "WACC")
    wacc_row = row
    n_seg = len(data["segments"])
    wacc_selected_assumptions = 6 + n_seg * 8 + 1 + 5 + 1 + 5 + 1 + 5 + 1 + 1 + 1 + 1 + 1 + 1 + 4
    c = _col(first_data_col)
    ws.cell(row=row, column=first_data_col,
            value=f"=Assumptions!{c}{wacc_selected_assumptions}").number_format = PCT_FORMAT
    row += 1

    _set_label(ws, row, 1, "Terminal Value Multiple")
    tv_mult_row = row
    tv_selected_assumptions = wacc_selected_assumptions + 6
    ws.cell(row=row, column=first_data_col,
            value=f"=Assumptions!{c}{tv_selected_assumptions}").number_format = '0.0"x"'
    row += 2

    # Discount Factor
    _set_label(ws, row, 1, "Discount Factor")
    df_row = row
    wacc_cell = f"${_col(first_data_col)}${wacc_row}"
    for i in range(n_proj):
        val_col = proj_cols[i]
        formula = f"=1/(1+{wacc_cell})^{i+1}"
        ws.cell(row=row, column=val_col, value=formula).number_format = '0.0000'
    row += 1

    # PV of UFCF
    _set_label(ws, row, 1, "PV of UFCF ($M)")
    pv_ufcf_row = row
    for i in range(n_proj):
        val_col = proj_cols[i]
        c = _col(val_col)
        formula = f"={c}{ufcf_dcf_row}*{c}{df_row}"
        ws.cell(row=row, column=val_col, value=formula).number_format = NUM_FORMAT_M
    row += 2

    # Terminal Value (EV/EBITDA exit multiple — SECTOR_TV son múltiplos de EBITDA, no UFCF)
    _set_label(ws, row, 1, "Terminal Value ($M)", bold=True)
    tv_row = row
    last_proj = _col(max(proj_cols))
    tv_cell = f"${_col(first_data_col)}${tv_mult_row}"
    ws.cell(row=row, column=max(proj_cols),
            value=f"={last_proj}{ebitda_dcf_row}*{tv_cell}").number_format = NUM_FORMAT_M
    row += 1

    # PV of TV
    _set_label(ws, row, 1, "PV of Terminal Value ($M)")
    pv_tv_row = row
    ws.cell(row=row, column=max(proj_cols),
            value=f"={last_proj}{tv_row}*{last_proj}{df_row}").number_format = NUM_FORMAT_M
    row += 2

    # VALUATION SUMMARY
    _set_header(ws, row, 1, last_proj_col, "Valuation Summary")
    row += 1

    _set_label(ws, row, 1, "Sum of PV of UFCF ($M)")
    sum_pv_row = row
    pv_range = f"{_col(min(proj_cols))}{pv_ufcf_row}:{_col(max(proj_cols))}{pv_ufcf_row}"
    ws.cell(row=row, column=first_data_col, value=f"=SUM({pv_range})").number_format = NUM_FORMAT_M
    row += 1

    _set_label(ws, row, 1, "PV of Terminal Value ($M)")
    pv_tv_summary_row = row
    ws.cell(row=row, column=first_data_col,
            value=f"={_col(max(proj_cols))}{pv_tv_row}").number_format = NUM_FORMAT_M
    row += 1

    _set_label(ws, row, 1, "Enterprise Value ($M)", bold=True)
    ev_row = row
    c = _col(first_data_col)
    ws.cell(row=row, column=first_data_col,
            value=f"={c}{sum_pv_row}+{c}{pv_tv_summary_row}").number_format = NUM_FORMAT_M
    row += 2

    # Net Debt
    net_debt = 0
    try:
        bs = data.get("balance_sheet")
        if bs is not None and not bs.empty:
            latest_col = bs.columns[-1]
            cash_val = 0
            debt_val = 0
            for k in ["CashAndCashEquivalents", "Cash And Cash Equivalents", "CashCashEquivalentsAndShortTermInvestments"]:
                if k in bs.index:
                    v = bs.loc[k, latest_col]
                    if v is not None and not (isinstance(v, float) and np.isnan(v)):
                        cash_val = float(v)
                        break
            for k in ["TotalDebt", "Total Debt", "LongTermDebt"]:
                if k in bs.index:
                    v = bs.loc[k, latest_col]
                    if v is not None and not (isinstance(v, float) and np.isnan(v)):
                        debt_val = float(v)
                        break
            net_debt = debt_val - cash_val
    except Exception:
        pass

    # Ajustar net debt si hay banco cautivo (Financial Services)
    captive = metrics.get("captive_finance") if isinstance(metrics, dict) else None
    if captive and captive.get("detected"):
        industrial_debt = captive["estimated_industrial_debt"]
        cash_for_calc = cash_val if cash_val else 0
        net_debt = industrial_debt - cash_for_calc
        _set_label(ws, row, 1, "(-) Net Debt ($M) [ajustada — excl. Financial Services]")
    else:
        _set_label(ws, row, 1, "(-) Net Debt ($M)")
    net_debt_row = row
    _set_input(ws, row, first_data_col, -net_debt / 1e6)
    ws.cell(row=row, column=first_data_col).number_format = NUM_FORMAT_M
    row += 1

    _set_label(ws, row, 1, "Equity Value ($M)", bold=True)
    equity_row = row
    c = _col(first_data_col)
    ws.cell(row=row, column=first_data_col,
            value=f"={c}{ev_row}+{c}{net_debt_row}").number_format = NUM_FORMAT_M
    row += 2

    _set_label(ws, row, 1, "Shares Outstanding (M)")
    shares_row = row
    _set_input(ws, row, first_data_col, shares / 1e6 if shares > 1e3 else shares)
    ws.cell(row=row, column=first_data_col).number_format = NUM_FORMAT_M
    row += 1

    _set_label(ws, row, 1, "Equity Value per Share ($)", bold=True)
    evps_row = row
    c = _col(first_data_col)
    cell = ws.cell(row=row, column=first_data_col, value=f"={c}{equity_row}/{c}{shares_row}")
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(name="Calibri", size=14, bold=True, color="FF006600")
    row += 1

    _set_label(ws, row, 1, "Upside / Downside vs Current", bold=True)
    c = _col(first_data_col)
    cell = ws.cell(row=row, column=first_data_col, value=f"={c}{evps_row}/{current_price}-1")
    cell.number_format = PCT_FORMAT
    cell.font = Font(name="Calibri", size=12, bold=True)
    row += 3

    # SENSITIVITY TABLE
    _set_header(ws, row, 1, first_data_col + 8,
                "Sensitivity Analysis: WACC vs Terminal Value Multiple")
    row += 1

    base_tv = base_sc["terminal_multiple"]
    tv_range = [base_tv - 4, base_tv - 2, base_tv, base_tv + 2, base_tv + 4]

    _set_label(ws, row, 1, "WACC \\ TV Multiple")
    for j, tv in enumerate(tv_range):
        col = first_data_col + j
        cell = ws.cell(row=row, column=col, value=f"{tv:.0f}x")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    sens_header_row = row
    row += 1

    base_wacc = base_sc["wacc"]
    wacc_range = [base_wacc + d for d in [-0.03, -0.02, -0.01, 0, 0.01, 0.02, 0.03]]

    for i, wacc_val in enumerate(wacc_range):
        sens_row = row + i
        cell = ws.cell(row=sens_row, column=1, value=wacc_val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.number_format = PCT_FORMAT

        for j, tv_val in enumerate(tv_range):
            col = first_data_col + j

            parts = []
            for k in range(n_proj):
                ufcf_cell = f"{_col(proj_cols[k])}${ufcf_dcf_row}"
                parts.append(f"{ufcf_cell}/(1+$A${sens_row})^{k+1}")

            last_ebitda = f"{_col(max(proj_cols))}${ebitda_dcf_row}"
            tv_formula = f"{last_ebitda}*{tv_val}/(1+$A${sens_row})^{n_proj}"

            nd_cell = f"${_col(first_data_col)}${net_debt_row}"
            sh_cell = f"${_col(first_data_col)}${shares_row}"

            sum_parts = "+".join(parts)
            full_formula = f"=({sum_parts}+{tv_formula}+{nd_cell})/{sh_cell}"

            cell = ws.cell(row=sens_row, column=col, value=full_formula)
            cell.number_format = CURRENCY_FORMAT

            if abs(wacc_val - base_wacc) < 0.001 and abs(tv_val - base_tv) < 0.1:
                cell.fill = INPUT_FILL
                cell.font = Font(name="Calibri", size=10, bold=True)

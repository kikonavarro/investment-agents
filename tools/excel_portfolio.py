"""
CRUD del archivo Excel de cartera (mi_cartera.xlsx).
Tres hojas: Posiciones, Transacciones, Watchlist.
Python hace todo el trabajo — Claude solo recibe el resumen formateado.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from datetime import date, datetime
from pathlib import Path
from config.settings import PORTFOLIO_FILE


# ─── Constantes de columnas (hoja Posiciones) ─────────────────────────────────
# A=Ticker B=Nombre C=Tipo D=Source E=Shares F=AvgPrice G=BuyDate
# H=CurrentPrice I=Target J=StopLoss K=Notas
# Para fondos manuales: E=InvestedAmount F=CurrentValue G=LastUpdated

COL = {
    "ticker": 1, "name": 2, "type": 3, "source": 4,
    "shares": 5, "avg_price": 6, "buy_date": 7,
    "current_price": 8, "target": 9, "stop_loss": 10, "notes": 11,
    # Fondos manuales (reutilizan columnas E y F)
    "invested_amount": 5, "current_value": 6, "last_updated": 7,
}


# ─── Inicialización ────────────────────────────────────────────────────────────

def create_portfolio_file():
    """Crea el archivo Excel con estructura inicial si no existe."""
    if PORTFOLIO_FILE.exists():
        return

    PORTFOLIO_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Hoja 1: Posiciones
    ws1 = wb.active
    ws1.title = "Posiciones"
    headers = [
        "Ticker", "Nombre", "Tipo", "Source",
        "Shares / Inv.Amount", "Avg Price / Cur.Value", "Buy Date / Last Updated",
        "Current Price", "Target", "Stop Loss", "Notas"
    ]
    ws1.append(headers)
    _style_header_row(ws1, 1)

    # Hoja 2: Transacciones
    ws2 = wb.create_sheet("Transacciones")
    ws2.append(["Fecha", "Ticker/Nombre", "Tipo", "Acción", "Shares/Cantidad", "Precio", "Comisión", "Notas"])
    _style_header_row(ws2, 1)

    # Hoja 3: Watchlist
    ws3 = wb.create_sheet("Watchlist")
    ws3.append(["Ticker", "Nombre", "Target Buy Price", "Notas", "Añadido"])
    _style_header_row(ws3, 1)

    wb.save(PORTFOLIO_FILE)
    print(f"Cartera creada en: {PORTFOLIO_FILE}")


# ─── Lectura ────────────────────────────────────────────────────────────────────

def read_portfolio() -> list[dict]:
    """Lee todas las posiciones. Devuelve lista de dicts."""
    _ensure_file()
    wb = openpyxl.load_workbook(PORTFOLIO_FILE, data_only=True)
    ws = wb["Posiciones"]
    positions = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] and not row[1]:  # Fila vacía
            continue
        source = row[COL["source"] - 1] or "auto"

        if source == "manual":
            pos = {
                "ticker": None,
                "name": row[COL["name"] - 1],
                "type": row[COL["type"] - 1] or "fund",
                "source": "manual",
                "invested_amount": _to_float(row[COL["invested_amount"] - 1]),
                "current_value": _to_float(row[COL["current_value"] - 1]),
                "last_updated": str(row[COL["last_updated"] - 1] or ""),
                "notes": row[COL["notes"] - 1],
            }
        else:
            pos = {
                "ticker": row[COL["ticker"] - 1],
                "name": row[COL["name"] - 1],
                "type": row[COL["type"] - 1] or "stock",
                "source": "auto",
                "shares": _to_float(row[COL["shares"] - 1]),
                "avg_price": _to_float(row[COL["avg_price"] - 1]),
                "buy_date": str(row[COL["buy_date"] - 1] or ""),
                "current_price": _to_float(row[COL["current_price"] - 1]),
                "target": _to_float(row[COL["target"] - 1]),
                "stop_loss": _to_float(row[COL["stop_loss"] - 1]),
                "notes": row[COL["notes"] - 1],
            }
        positions.append(pos)

    return positions


def read_watchlist() -> list[dict]:
    """Lee la watchlist."""
    _ensure_file()
    wb = openpyxl.load_workbook(PORTFOLIO_FILE, data_only=True)
    ws = wb["Watchlist"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        items.append({
            "ticker": row[0],
            "name": row[1],
            "target_buy": _to_float(row[2]),
            "notes": row[3],
            "added": str(row[4] or ""),
        })
    return items


# ─── Escritura ──────────────────────────────────────────────────────────────────

def update_prices(price_map: dict[str, float]):
    """Actualiza precios de posiciones automáticas (source=auto)."""
    _ensure_file()
    wb = openpyxl.load_workbook(PORTFOLIO_FILE)
    ws = wb["Posiciones"]

    for row in ws.iter_rows(min_row=2):
        ticker = row[COL["ticker"] - 1].value
        source = row[COL["source"] - 1].value
        if source == "auto" and ticker and ticker in price_map:
            row[COL["current_price"] - 1].value = price_map[ticker]

    wb.save(PORTFOLIO_FILE)


def update_manual_position(name: str, current_value: float):
    """Actualiza valor de un fondo manual (source=manual)."""
    _ensure_file()
    wb = openpyxl.load_workbook(PORTFOLIO_FILE)
    ws = wb["Posiciones"]
    today = date.today().isoformat()

    for row in ws.iter_rows(min_row=2):
        if row[COL["name"] - 1].value == name and row[COL["source"] - 1].value == "manual":
            row[COL["current_value"] - 1].value = current_value
            row[COL["last_updated"] - 1].value = today
            break

    wb.save(PORTFOLIO_FILE)


def add_position(
    name: str,
    source: str,                   # "auto" | "manual"
    position_type: str = "stock",  # "stock" | "fund"
    # Auto
    ticker: str = None,
    shares: float = None,
    avg_price: float = None,
    buy_date: str = None,
    current_price: float = None,
    target: float = None,
    stop_loss: float = None,
    # Manual
    invested_amount: float = None,
    current_value: float = None,
    notes: str = None,
):
    """Añade una nueva posición a la hoja Posiciones."""
    _ensure_file()
    wb = openpyxl.load_workbook(PORTFOLIO_FILE)
    ws = wb["Posiciones"]

    if source == "auto":
        row = [ticker, name, position_type, source,
               shares, avg_price, buy_date or date.today().isoformat(),
               current_price, target, stop_loss, notes]
    else:
        row = [None, name, position_type, source,
               invested_amount, current_value, date.today().isoformat(),
               None, None, None, notes]

    ws.append(row)
    wb.save(PORTFOLIO_FILE)


def add_transaction(
    ticker_or_name: str,
    action: str,         # "buy" | "sell"
    amount: float,       # shares para stocks, € para fondos
    price: float = None,
    position_type: str = "stock",
    fee: float = 0.0,
    notes: str = None,
):
    """Añade una transacción al historial."""
    _ensure_file()
    wb = openpyxl.load_workbook(PORTFOLIO_FILE)
    ws = wb["Transacciones"]
    ws.append([
        date.today().isoformat(), ticker_or_name, position_type,
        action, amount, price, fee, notes or ""
    ])
    wb.save(PORTFOLIO_FILE)


def add_to_watchlist(ticker: str, name: str, target_buy: float = None, notes: str = None):
    """Añade una empresa a la watchlist."""
    _ensure_file()
    wb = openpyxl.load_workbook(PORTFOLIO_FILE)
    ws = wb["Watchlist"]
    ws.append([ticker, name, target_buy, notes or "", date.today().isoformat()])
    wb.save(PORTFOLIO_FILE)


# ─── Cálculo de resumen (Python puro) ──────────────────────────────────────────

def get_portfolio_summary() -> dict:
    """
    Calcula P&L y métricas de cartera. Todo en Python, sin Claude.
    Este dict es lo que se pasa al formatter antes de enviar a Claude.
    """
    positions = read_portfolio()
    if not positions:
        return {"total_value": 0, "total_invested": 0, "total_pnl_pct": 0, "positions": []}

    total_value = 0
    total_invested = 0
    summary_positions = []

    for pos in positions:
        if pos["source"] == "auto":
            cp = pos.get("current_price") or 0
            ap = pos.get("avg_price") or 0
            sh = pos.get("shares") or 0
            current_val = cp * sh
            invested_val = ap * sh
            pnl_pct = ((cp - ap) / ap * 100) if ap > 0 else 0
            vs_target = None
            if pos.get("target") and cp > 0:
                vs_target = round((pos["target"] - cp) / cp * 100, 1)
            alert = None
            if pos.get("target") and cp >= pos["target"]:
                alert = "TARGET ALCANZADO"
            elif pos.get("stop_loss") and cp <= pos["stop_loss"]:
                alert = "STOP LOSS"
        else:
            current_val = pos.get("current_value") or 0
            invested_val = pos.get("invested_amount") or 0
            pnl_pct = ((current_val - invested_val) / invested_val * 100) if invested_val > 0 else 0
            vs_target = None
            alert = None
            # Detectar si lleva más de 30 días sin actualizar
            last = pos.get("last_updated", "")
            if last:
                try:
                    days_old = (date.today() - date.fromisoformat(last[:10])).days
                    if days_old > 30:
                        alert = f"SIN ACTUALIZAR ({days_old}d)"
                except ValueError:
                    pass

        total_value += current_val
        total_invested += invested_val

        summary_positions.append({
            "name": pos.get("ticker") or pos.get("name"),
            "source": pos["source"],
            "invested": round(invested_val, 2),
            "current": round(current_val, 2),
            "pnl_pct": round(pnl_pct, 2),
            "vs_target": vs_target,
            "alert": alert,
            "needs_update": pos["source"] == "manual" and alert and "SIN ACTUALIZAR" in alert,
        })

    total_pnl = ((total_value - total_invested) / total_invested * 100) if total_invested > 0 else 0

    return {
        "total_value": round(total_value, 2),
        "total_invested": round(total_invested, 2),
        "total_pnl_pct": round(total_pnl, 2),
        "positions": summary_positions,
        "alerts": [p for p in summary_positions if p.get("alert")],
    }


# ─── Helpers ────────────────────────────────────────────────────────────────────

def _ensure_file():
    if not PORTFOLIO_FILE.exists():
        create_portfolio_file()


def _to_float(val) -> float | None:
    try:
        return float(val) if val is not None else None
    except (TypeError, ValueError):
        return None


def _style_header_row(ws, row_num: int):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
